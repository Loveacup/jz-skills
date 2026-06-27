#!/usr/bin/env python3
"""
DingTalk 统一媒体管线 — 图片 + 文件下载与内容提取
输入：dingwave 解密的 SQLite DB 路径
输出：下载原文件 + 文字提取 → JSON（供 monitor 脚本消费）
"""
import sys, os, json, sqlite3, base64, subprocess, hashlib, tempfile
from datetime import datetime

CLASS_CID = "52993580719"
DDSID = "k0_74fa0b0b9ed5236ac15b_0b0b74fa6a23d59e691c001a5901b804caabf67c80e6"
DT_S = "u-25adff5-9e9bfa7d63-21084bdf-5438b1-610a7ac9-a2d6775a-7931-48b3-a751-795340e045c0"

DATA_DIR = os.path.expanduser("~/.hermes/data/dingtalk_media")
CACHE_FILE = os.path.join(DATA_DIR, ".pipeline_cache.json")
STATE_FILE = os.path.join(DATA_DIR, ".last_processed_id")

VISION_MODEL = "gpt-5.5"

os.makedirs(DATA_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════
# msgpack decoder (for authMediaId)
# ══════════════════════════════════════════════════════════
def msgpack_decode(b):
    pos = 0
    def rd():
        nonlocal pos
        c = b[pos]; pos += 1
        if c <= 0x7f: return c
        if 0x80 <= c <= 0x8f:
            n = c & 0x0f; d = {}
            for _ in range(n): k = rd(); v = rd(); d[k] = v
            return d
        if 0x90 <= c <= 0x9f: return [rd() for _ in range(c & 0x0f)]
        if 0xa0 <= c <= 0xbf:
            n = c & 0x1f; s = b[pos:pos+n]; pos += n
            try: return s.decode('utf-8')
            except: return s.decode('latin-1')
        if c == 0xcc: v = b[pos]; pos += 1; return v
        if c == 0xcd: v = int.from_bytes(b[pos:pos+2], 'big'); pos += 2; return v
        if c == 0xce: v = int.from_bytes(b[pos:pos+4], 'big'); pos += 4; return v
        if c == 0xcf: v = int.from_bytes(b[pos:pos+8], 'big'); pos += 8; return v
        if c == 0xd0: v = int.from_bytes(b[pos:pos+1], 'big', signed=True); pos += 1; return v
        if c == 0xd1: v = int.from_bytes(b[pos:pos+2], 'big', signed=True); pos += 2; return v
        if c == 0xd2: v = int.from_bytes(b[pos:pos+4], 'big', signed=True); pos += 4; return v
        if c == 0xd3: v = int.from_bytes(b[pos:pos+8], 'big', signed=True); pos += 8; return v
        if 0xe0 <= c <= 0xff: return c - 256
    return rd()

def build_image_url(amid):
    body = amid[1:]
    raw = base64.b64decode(body.replace('-','+').replace('_','/') + '=' * (-len(body) % 4))
    obj = msgpack_decode(raw)
    typ = obj.get(2, 'jpg')
    return f"https://down.dingtalk.com/ddmedia/{body}.{typ}", obj

# ══════════════════════════════════════════════════════════
# Extract media from DB
# ══════════════════════════════════════════════════════════
def extract_media(db_path, last_id=0):
    conn = sqlite3.connect(db_path)
    rows = conn.execute("""
        SELECT rowid, datetime(created_at/1000, 'unixepoch', '+8 hours'), content_json
        FROM messages WHERE cid=? ORDER BY created_at ASC
    """, (CLASS_CID,)).fetchall()
    conn.close()

    images = []
    files = []
    for rowid, ts, raw in rows:
        if rowid <= last_id:
            continue
        d = json.loads(raw)
        for att in d.get('attachments', []):
            ext_s = att.get('extension', '{}')
            ext = json.loads(ext_s) if ext_s else {}
            att_type = att.get('type', 0)

            # ── File attachments (type=0, has f_id) ──
            if att_type == 0 and ext.get('f_id'):
                files.append({
                    'rowid': rowid, 'time': ts,
                    'f_id': ext['f_id'],
                    's_id': ext.get('s_id', ''),
                    'f_name': ext.get('f_name', 'unknown'),
                    'f_size': ext.get('f_size', 0),
                })
                continue

            # ── Rich text (type=3100) → images ──
            if att_type != 3100:
                continue
            p2_s = ext.get('payloadV2', '{}')
            if not p2_s: continue
            p2 = json.loads(p2_s) if isinstance(p2_s, str) else p2_s
            for c in p2.get('contents', []):
                if c.get('type') != 'markdown': continue
                for item in c.get('text', {}).get('items', []):
                    if item.get('type') != 'image': continue
                    data = item.get('data', {})
                    auth = data.get('authMediaId', '')
                    if not auth: continue
                    w = item.get('style', {}).get('width', 0)
                    h = item.get('style', {}).get('height', 0)
                    images.append({
                        'rowid': rowid, 'time': ts,
                        'authMediaId': auth, 'width': w, 'height': h,
                    })
    return images, files

# ══════════════════════════════════════════════════════════
# Download
# ══════════════════════════════════════════════════════════
def download(url, out_path, cookie_str, timeout=60):
    result = subprocess.run([
        'curl', '-s', '-f', '-L', '-o', out_path,
        '-b', cookie_str,
        '-H', 'User-Agent: Mozilla/5.0',
        url
    ], capture_output=True, timeout=timeout)
    if result.returncode != 0:
        return f"curl failed (exit {result.returncode})"
    size = os.path.getsize(out_path)
    if size < 100:
        return f"too small ({size} bytes, likely error page)"
    return None

# ══════════════════════════════════════════════════════════
# Content extraction
# ══════════════════════════════════════════════════════════
def ocr_image(path):
    """Vision model OCR for images"""
    import urllib.request, ssl
    with open(path, 'rb') as f:
        img_b64 = base64.b64encode(f.read()).decode()

    api_key = os.environ.get('OPENAI_API_KEY', '')
    api_base = os.environ.get('OPENAI_BASE_URL', 'https://api.openai.com/v1')
    if not api_key:
        try:
            import yaml
            for prof in ['default', 'cron-worker']:
                cfg_path = os.path.expanduser(f'~/.hermes/profiles/{prof}/config.yaml')
                if os.path.exists(cfg_path):
                    with open(cfg_path) as f:
                        cfg = yaml.safe_load(f)
                    for name, p in cfg.get('providers', {}).items():
                        if 'openai' in p.get('base_url', ''):
                            api_key = p.get('api_key', '')
                            api_base = p.get('base_url', api_base)
                            break
        except: pass

    payload = {
        'model': VISION_MODEL,
        'messages': [{'role': 'user', 'content': [
            {'type': 'text', 'text': '请提取这张图片中的所有文字内容，保持原始格式和排版。只输出文字，不要额外说明。'},
            {'type': 'image_url', 'image_url': {'url': f'data:image/jpeg;base64,{img_b64}'}}
        ]}],
        'max_tokens': 2000
    }
    req = urllib.request.Request(f'{api_base}/chat/completions',
        data=json.dumps(payload).encode(),
        headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}'})
    try:
        resp = urllib.request.urlopen(req, context=ssl.create_default_context(), timeout=60)
        return json.loads(resp.read())['choices'][0]['message']['content']
    except Exception as e:
        return f"[OCR 失败: {e}]"

def extract_pdf(path):
    """Extract text from PDF"""
    try:
        import fitz
        doc = fitz.open(path)
        texts = []
        for page in doc:
            t = page.get_text()
            if t.strip():
                texts.append(t.strip())
        doc.close()
        return '\n'.join(texts) if texts else "[PDF 无可提取文字]"
    except Exception as e:
        return f"[PDF 提取失败: {e}]"

def extract_excel(path, ext):
    """Extract text from Excel"""
    try:
        if ext.lower() in ('.xls',):
            import xlrd
            wb = xlrd.open_workbook(path)
            texts = []
            for sheet in wb.sheets():
                for row in range(sheet.nrows):
                    row_text = ' | '.join(
                        str(sheet.cell_value(row, col)) for col in range(sheet.ncols)
                        if sheet.cell_value(row, col) != ''
                    )
                    if row_text.strip():
                        texts.append(row_text)
            return '\n'.join(texts) if texts else "[Excel 无内容]"
        elif ext.lower() in ('.xlsx',):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                texts = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    texts.append(f'--- {name} ---')
                    for row in ws.iter_rows(values_only=True):
                        row_text = ' | '.join(str(c) for c in row if c is not None)
                        if row_text.strip():
                            texts.append(row_text)
                return '\n'.join(texts) if texts else "[Excel 无内容]"
            except ImportError:
                return "[需安装 openpyxl 以读取 .xlsx]"
        return f"[不支持的格式: {ext}]"
    except Exception as e:
        return f"[Excel 提取失败: {e}]"

# ══════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════
def main():
    if len(sys.argv) < 2:
        print("Usage: dingtalk-media-pipeline.py <merged_db_path>", file=sys.stderr)
        sys.exit(1)

    db_path = sys.argv[1]

    # Load cache + state
    cache = {}
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE) as f:
            cache = json.load(f)

    last_id = 0
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as f:
            last_id = int(f.read().strip() or 0)

    images, files = extract_media(db_path, last_id)
    outputs = []
    max_rowid = last_id
    print(f"Found {len(images)} images, {len(files)} files")

    # ── Process images ──
    for img in images:
        aid = img['authMediaId']
        img_hash = hashlib.md5(aid.encode()).hexdigest()[:12]
        safe_name = f"img_{img['time'][:10]}_{img_hash}.jpg"
        out_path = os.path.join(DATA_DIR, safe_name)
        max_rowid = max(max_rowid, img['rowid'])

        if img_hash in cache:
            text = cache[img_hash]
        elif os.path.exists(out_path):
            text = ocr_image(out_path)
            cache[img_hash] = text
        else:
            url, meta = build_image_url(aid)
            err = download(url, out_path, f"dd_sid={DDSID}")
            if err:
                print(f"  [skip img] {img['time']}: {err}")
                continue
            print(f"  [img] {img['time']} {img['width']}x{img['height']} → {safe_name}")
            text = ocr_image(out_path)
            cache[img_hash] = text

        outputs.append({
            'type': 'image', 'time': img['time'],
            'file': safe_name,
            'size': f"{img['width']}x{img['height']}",
            'text': text,
        })

    # ── Process files ──
    for f in files:
        fid = f['f_id']
        file_hash = hashlib.md5(f"{f['s_id']}_{fid}".encode()).hexdigest()[:12]
        ext = os.path.splitext(f['f_name'])[1].lower()
        safe_name = f"file_{f['time'][:10]}_{file_hash}{ext}"
        out_path = os.path.join(DATA_DIR, safe_name)
        max_rowid = max(max_rowid, f['rowid'])

        if file_hash in cache:
            text = cache[file_hash]
        elif os.path.exists(out_path):
            # Re-extract text from cached file
            if ext == '.pdf':
                text = extract_pdf(out_path)
            elif ext in ('.xls', '.xlsx'):
                text = extract_excel(out_path, ext)
            else:
                text = f"[{f['f_name']} — 已下载，格式 {ext}]"
            cache[file_hash] = text
        else:
            url = f"https://space.dingtalk.com/attachment/mdown?biztype=file&bizid={f['s_id']}&objectid={fid}"
            err = download(url, out_path, f"dt_s={DT_S}")
            if err:
                print(f"  [skip file] {f['time']} {f['f_name']}: {err}")
                continue
            print(f"  [file] {f['time']} {f['f_name']} ({f['f_size']}B) → {safe_name}")

            if ext == '.pdf':
                text = extract_pdf(out_path)
            elif ext in ('.xls', '.xlsx'):
                text = extract_excel(out_path, ext)
            else:
                text = f"[{f['f_name']} — {os.path.getsize(out_path)}B]"
            cache[file_hash] = text

        outputs.append({
            'type': 'file', 'time': f['time'],
            'file': safe_name,
            'name': f['f_name'],
            'size': f['f_size'],
            'text': text,
        })

    # Save state
    with open(CACHE_FILE, 'w') as f:
        json.dump(cache, f, ensure_ascii=False)
    with open(STATE_FILE, 'w') as f:
        f.write(str(max_rowid))

    print(json.dumps(outputs, ensure_ascii=False, indent=2))
    print(f"\nDone: {len(outputs)} items processed")

if __name__ == '__main__':
    main()
